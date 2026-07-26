"""
期刊论文引用通知自动化管理系统 - Flask Web 应用
"""

import io
import json
import os
import secrets
import threading
from pathlib import Path

import yaml
import pandas as pd
from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, jsonify, send_from_directory, send_file, session,
)
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from extract_references import process_input_directory, save_references_csv, split_reference_rows
from lookup_authors import lookup_emails, save_results, merge_manual_supplement
from send_emails import load_sent_log, prepare_email_data, render_email, repair_citing_authors_from_input, send_emails

BASE_DIR = Path(__file__).parent
RUNTIME_DIR = Path(os.environ.get("CITATION_NOTIFIER_HOME", BASE_DIR)).expanduser()
CONFIG_PATH = RUNTIME_DIR / "config.yaml"
DATA_DIR = RUNTIME_DIR / "data"
INPUT_DIR = DATA_DIR / "input"
PREVIEW_DIR = DATA_DIR / "preview"

ALLOWED_EXTENSIONS = {".doc", ".docx", ".pdf"}

JOURNALS = {
    "innovation": {
        "id": "innovation",
        "short_name": "双创",
        "name": "创新与创业教育",
        "name_en": "Innovation and Entrepreneurship Education",
        "show_name_en_in_email": False,
        "theme": "blue",
        "primary": "#1a5276",
        "primary_light": "#2980b9",
        "bg_light": "#f4f6f9",
        "mail_bg": "#f4f7fb",
        "data_dir": "data/innovation",
        "sender_name": "《创新与创业教育》编辑部",
        "editorial_office": "《创新与创业教育》编辑部",
        "intro": "《创新与创业教育》是教育部主管、中南大学主办的学术期刊，创刊于2010年，为中国人文社会科学 AMI 综合评价 A 刊核心期刊扩展版，被 CNKI、万方数据、维普资讯等收录。本刊常设“学术前沿”“教学研究”“教育治理”“数智教育”“劳动教育”“创业管理”“双创实践”“技术创新”“青年论坛”等栏目，致力于打造集学术引领、实践创新与政策研讨于一体的高水平学术交流平台。",
        "attachment_note": "以上引用信息供您知悉与参考，具体内容请以本刊正式出版版本为准。",
    },
    "csu_social": {
        "id": "csu_social",
        "short_name": "社科",
        "name": "中南大学学报（社会科学版）",
        "name_en": "Journal of Central South University (Social Science)",
        "show_name_en_in_email": False,
        "theme": "brown",
        "primary": "#6f4e37",
        "primary_light": "#8a664a",
        "bg_light": "#f7f3ee",
        "mail_bg": "#f8f1e9",
        "data_dir": "data/csu_social",
        "sender_name": "《中南大学学报（社会科学版）》编辑部",
        "editorial_office": "《中南大学学报（社会科学版）》编辑部",
        "intro": "《中南大学学报（社会科学版）》是由教育部主管、中南大学主办的综合性学术理论期刊，是 CSSCI 来源期刊、中文核心期刊、中国人文社会科学 AMI 综合评价 A 刊核心期刊、全国高校权威社科期刊、湖南省社会科学基金资助期刊，设有马克思主义、哲学、法学、经济学、管理学、文学、政治学与社会学等栏目。",
        "attachment_note": "以上引用信息供您知悉与参考，具体内容请以本刊正式出版版本为准。",
    },
}

# ---- 全局任务状态 ----
task_status = {
    "running": False,
    "stage": "",
    "progress": "",
    "log": [],
}


def load_config() -> dict:
    config_path = runtime_path("config.yaml")
    if config_path.exists():
        with open(config_path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    return {}


def save_config(config: dict):
    config_path = runtime_path("config.yaml")
    config_path.parent.mkdir(parents=True, exist_ok=True)
    with open(config_path, "w", encoding="utf-8") as f:
        yaml.dump(config, f, allow_unicode=True, default_flow_style=False, sort_keys=False)


def runtime_path(relative_path: str | Path) -> Path:
    path = Path(relative_path)
    if path.is_absolute():
        return path
    return RUNTIME_DIR / path


def app_path(relative_path: str | Path) -> Path:
    path = Path(relative_path)
    if path.is_absolute():
        return path
    return BASE_DIR / path


def get_active_journal_id() -> str:
    requested = request.args.get("journal", "").strip()
    if requested in JOURNALS:
        session["active_journal"] = requested
        return requested
    saved = session.get("active_journal", "innovation")
    if saved in JOURNALS:
        return saved
    session["active_journal"] = "innovation"
    return "innovation"


def get_active_journal() -> dict:
    return JOURNALS[get_active_journal_id()]


def build_journal_paths(journal: dict, config: dict) -> dict:
    data_root = journal["data_dir"]
    base_paths = config.get("paths", {})
    return {
        "input_dir": f"{data_root}/input",
        "references_csv": f"{data_root}/references.csv",
        "other_references_csv": f"{data_root}/other_references.csv",
        "references_no_author_csv": f"{data_root}/other_references.csv",
        "authors_emails_csv": f"{data_root}/authors_emails.csv",
        "manual_supplement_csv": f"{data_root}/manual_supplement.csv",
        "failed_lookup_csv": f"{data_root}/failed_lookup.csv",
        "sent_log_csv": f"{data_root}/sent_log.csv",
        "preview_dir": f"{data_root}/preview",
        "email_overrides_json": f"{data_root}/email_overrides.json",
        "local_db_dir": base_paths.get("local_db_dir", "数据"),
    }


def get_runtime_config() -> dict:
    config = load_config()
    journal = get_active_journal()
    config["paths"] = build_journal_paths(journal, config)
    config["journal"] = {
        "name": journal["name"],
        "name_en": journal["name_en"],
        "show_name_en_in_email": journal.get("show_name_en_in_email", True),
        "intro": journal["intro"],
        "website": config.get("journal", {}).get("website", ""),
        "theme_color": journal["primary"],
        "theme_bg": journal["mail_bg"],
        "editorial_office": journal["editorial_office"],
        "attachment_note": journal["attachment_note"],
    }
    config.setdefault("smtp", {})
    config["smtp"]["sender_name"] = journal["sender_name"]
    return config


def active_path(path_key: str) -> Path:
    config = get_runtime_config()
    return runtime_path(config["paths"][path_key])


def ensure_journal_dirs(paths: dict):
    for key in ("input_dir", "preview_dir"):
        runtime_path(paths[key]).mkdir(parents=True, exist_ok=True)


def split_existing_references(paths: dict):
    """Normalize old mixed reference files into papers and other references."""
    ref_csv = runtime_path(paths["references_csv"])
    other_csv = runtime_path(paths["other_references_csv"])
    if not ref_csv.exists():
        return

    try:
        df_ref = pd.read_csv(ref_csv, encoding="utf-8-sig").fillna("")
    except Exception:
        return

    rows = df_ref.to_dict("records")
    papers, other_from_ref = split_reference_rows(rows)
    changed = len(papers) != len(rows)

    if other_csv.exists():
        try:
            df_other = pd.read_csv(other_csv, encoding="utf-8-sig").fillna("")
            existing_other = df_other.to_dict("records")
        except Exception:
            existing_other = []
    else:
        existing_other = []

    if other_from_ref:
        seen = {
            (str(row.get("source_file", "")), str(row.get("raw_text", "")), str(row.get("title", "")))
            for row in existing_other
        }
        for row in other_from_ref:
            key = (str(row.get("source_file", "")), str(row.get("raw_text", "")), str(row.get("title", "")))
            if key not in seen:
                existing_other.append(row)
                seen.add(key)

    if changed:
        ref_csv.parent.mkdir(parents=True, exist_ok=True)
        pd.DataFrame(papers, columns=df_ref.columns).to_csv(ref_csv, index=False, encoding="utf-8-sig")
        if existing_other:
            other_csv.parent.mkdir(parents=True, exist_ok=True)
            pd.DataFrame(existing_other).to_csv(other_csv, index=False, encoding="utf-8-sig")


def resolve_secret_key() -> str:
    """优先使用环境变量或配置中的密钥，避免硬编码固定值。"""
    env_key = os.environ.get("CITATION_NOTIFIER_SECRET_KEY", "").strip()
    if env_key:
        return env_key

    config_key = str(load_config().get("app", {}).get("secret_key", "")).strip()
    if config_key:
        return config_key

    return secrets.token_hex(32)


app = Flask(__name__)
app.secret_key = resolve_secret_key()
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB


@app.context_processor
def inject_journal_context():
    current = get_active_journal()
    def journal_url(journal_id: str):
        endpoint = request.endpoint or "index"
        values = dict(request.view_args or {})
        values["journal"] = journal_id
        try:
            return url_for(endpoint, **values)
        except Exception:
            return url_for("index", journal=journal_id)
    return {
        "journals": JOURNALS,
        "current_journal": current,
        "journal_url": journal_url,
    }


def get_data_stats() -> dict:
    """获取各阶段数据统计"""
    stats = {
        "input_files": 0,
        "references": 0,
        "emails_found": 0,
        "emails_failed": 0,
        "emails_sent": 0,
    }

    config = get_runtime_config()
    paths = config.get("paths", {})
    input_dir = runtime_path(paths["input_dir"])

    # 输入文件数
    if input_dir.exists():
        stats["input_files"] = len([
            f for f in input_dir.iterdir()
            if f.suffix.lower() in ALLOWED_EXTENSIONS
        ])

    # 参考文献数
    ref_csv = runtime_path(paths.get("references_csv", "data/references.csv"))
    if ref_csv.exists():
        try:
            df = pd.read_csv(ref_csv, encoding="utf-8-sig").fillna("")
            stats["references"] = len(df)
        except Exception:
            pass

    # 邮箱检索结果
    emails_csv = runtime_path(paths.get("authors_emails_csv", "data/authors_emails.csv"))
    if emails_csv.exists():
        try:
            df = pd.read_csv(emails_csv, encoding="utf-8-sig").fillna("")
            stats["emails_found"] = len(df)
        except Exception:
            pass

    failed_csv = runtime_path(paths.get("failed_lookup_csv", "data/failed_lookup.csv"))
    if failed_csv.exists():
        try:
            df = pd.read_csv(failed_csv, encoding="utf-8-sig").fillna("")
            stats["emails_failed"] = len(df)
        except Exception:
            pass

    # 已发送
    sent_csv = runtime_path(paths.get("sent_log_csv", "data/sent_log.csv"))
    if sent_csv.exists():
        try:
            df = load_sent_log(str(sent_csv))
            stats["emails_sent"] = len(
                df[df["status"] == "成功"].drop_duplicates(subset=["timestamp", "email"])
            )
        except Exception:
            pass

    return stats


# ==================== 路由 ====================

@app.route("/")
def index():
    stats = get_data_stats()
    return render_template("index.html", stats=stats)


# ---- 配置页 ----

@app.route("/config", methods=["GET", "POST"])
def config_page():
    config = load_config()

    if request.method == "POST":
        # 从表单更新配置
        config.setdefault("smtp", {})
        config["smtp"]["server"] = request.form.get("smtp_server", "").strip()
        config["smtp"]["port"] = int(request.form.get("smtp_port", 465))
        config["smtp"]["use_ssl"] = request.form.get("smtp_ssl") == "on"
        config["smtp"]["username"] = request.form.get("smtp_username", "").strip()
        config["smtp"]["password"] = request.form.get("smtp_password", "").strip()
        config["smtp"]["sender_name"] = request.form.get("smtp_sender_name", "").strip()

        config.setdefault("journal", {})
        config["journal"]["name"] = request.form.get("journal_name", "").strip()
        config["journal"]["name_en"] = request.form.get("journal_name_en", "").strip()
        config["journal"]["website"] = request.form.get("journal_website", "").strip()

        config.setdefault("paths", {})
        config["paths"]["local_db_dir"] = request.form.get("local_db_dir", "数据").strip() or "数据"

        save_config(config)
        flash("配置已保存", "success")
        return redirect(url_for("config_page"))

    return render_template("config_page.html", config=get_runtime_config())


@app.route("/config/test-smtp", methods=["POST"])
def test_smtp():
    """测试 SMTP 连接"""
    import smtplib
    import ssl

    config = get_runtime_config()
    smtp = config.get("smtp", {})

    try:
        if smtp.get("use_ssl", True):
            context = ssl.create_default_context()
            server = smtplib.SMTP_SSL(smtp["server"], smtp["port"], context=context, timeout=10)
        else:
            server = smtplib.SMTP(smtp["server"], smtp["port"], timeout=10)
            server.starttls()

        server.login(smtp["username"], smtp["password"])
        server.quit()
        return jsonify({"success": True, "message": "SMTP 连接成功！"})
    except Exception as e:
        return jsonify({"success": False, "message": f"连接失败: {e}"})


# ---- 第一阶段：上传与提取 ----

@app.route("/extract")
def extract_page():
    config = get_runtime_config()
    paths = config.get("paths", {})
    input_dir = runtime_path(paths["input_dir"])
    ensure_journal_dirs(paths)
    split_existing_references(paths)

    files = []
    if input_dir.exists():
        for f in sorted(input_dir.iterdir()):
            if f.suffix.lower() in ALLOWED_EXTENSIONS:
                size_kb = f.stat().st_size / 1024
                files.append({"name": f.name, "size": f"{size_kb:.1f} KB", "type": f.suffix})

    ref_csv = runtime_path(paths.get("references_csv", "data/references.csv"))
    references = []
    if ref_csv.exists():
        try:
            references = pd.read_csv(ref_csv, encoding="utf-8-sig").fillna("").to_dict("records")
        except Exception:
            pass

    no_author_csv = runtime_path(paths.get("other_references_csv", "data/other_references.csv"))
    no_author_refs = []
    if no_author_csv.exists():
        try:
            no_author_refs = pd.read_csv(no_author_csv, encoding="utf-8-sig").fillna("").to_dict("records")
        except Exception:
            pass

    return render_template("extract.html", files=files, references=references, no_author_refs=no_author_refs, task_status=task_status)


@app.route("/extract/upload", methods=["POST"])
def upload_files():
    config = get_runtime_config()
    paths = config.get("paths", {})
    input_dir = runtime_path(paths["input_dir"])
    input_dir.mkdir(parents=True, exist_ok=True)

    uploaded = request.files.getlist("files")
    count = 0
    for f in uploaded:
        if f.filename:
            ext = Path(f.filename).suffix.lower()
            if ext in ALLOWED_EXTENSIONS:
                filename = secure_filename(f.filename)
                # 保留中文文件名
                if not filename or filename == ext:
                    filename = f.filename
                f.save(str(input_dir / filename))
                count += 1

    flash(f"已上传 {count} 个文件", "success")
    return redirect(url_for("extract_page"))


@app.route("/extract/delete/<filename>", methods=["POST"])
def delete_file(filename):
    filepath = active_path("input_dir") / filename
    if filepath.exists():
        filepath.unlink()
        flash(f"已删除: {filename}", "info")
    return redirect(url_for("extract_page"))


@app.route("/extract/run", methods=["POST"])
def run_extract():
    if task_status["running"]:
        flash("已有任务正在运行，请等待完成", "warning")
        return redirect(url_for("extract_page"))

    config = get_runtime_config()
    paths = config.get("paths", {})
    input_dir = runtime_path(paths["input_dir"])
    ref_csv = str(runtime_path(paths.get("references_csv", "data/references.csv")))
    other_csv = str(runtime_path(paths.get("other_references_csv", "data/other_references.csv")))

    def run_in_background():
        task_status["running"] = True
        task_status["stage"] = "参考文献提取"
        task_status["progress"] = "处理中..."
        task_status["log"] = []
        try:
            results = process_input_directory(str(input_dir))
            if not results:
                task_status["progress"] = "未找到任何可处理的文件"
                return
            papers, other = save_references_csv(results, ref_csv, other_csv)
            errors = [r for r in results if r.get("error")]
            msg = f"完成：{len(results)} 篇文章，{len(papers)} 条论文，{len(other)} 条其他文献"
            if errors:
                msg += f"，{len(errors)} 篇有警告"
            task_status["progress"] = msg
        except Exception as e:
            task_status["progress"] = f"出错: {e}"
        finally:
            task_status["running"] = False

    thread = threading.Thread(target=run_in_background, daemon=True)
    thread.start()

    flash("参考文献提取已开始（后台运行中），请稍候刷新页面查看结果", "info")
    return redirect(url_for("extract_page"))


# ---- 第二阶段：邮箱检索 ----

@app.route("/lookup")
def lookup_page():
    config = get_runtime_config()
    paths = config.get("paths", {})
    split_existing_references(paths)

    # 检索结果
    emails_data = []
    emails_csv = runtime_path(paths.get("authors_emails_csv", "data/authors_emails.csv"))
    if emails_csv.exists():
        try:
            df = pd.read_csv(emails_csv, encoding="utf-8-sig").fillna("")
            emails_data = df.to_dict("records")
        except Exception:
            pass

    # 失败记录
    failed_data = []
    failed_csv = runtime_path(paths.get("failed_lookup_csv", "data/failed_lookup.csv"))
    if failed_csv.exists():
        try:
            df = pd.read_csv(failed_csv, encoding="utf-8-sig").fillna("")
            failed_data = df.to_dict("records")
        except Exception:
            pass

    # 检查是否有参考文献数据
    ref_csv = runtime_path(paths.get("references_csv", "data/references.csv"))
    has_references = ref_csv.exists()

    return render_template(
        "lookup.html",
        emails_data=emails_data,
        failed_data=failed_data,
        has_references=has_references,
        task_status=task_status,
    )


@app.route("/lookup/run", methods=["POST"])
def run_lookup():
    if task_status["running"]:
        flash("已有任务正在运行，请等待完成", "warning")
        return redirect(url_for("lookup_page"))

    config = get_runtime_config()
    paths = config.get("paths", {})
    split_existing_references(paths)
    ref_csv = str(runtime_path(paths.get("references_csv", "data/references.csv")))

    if not os.path.exists(ref_csv):
        flash("请先执行第一阶段（提取参考文献）", "warning")
        return redirect(url_for("lookup_page"))

    # 在后台线程中运行检索（避免超时）
    def run_in_background():
        task_status["running"] = True
        task_status["stage"] = "邮箱检索"
        task_status["log"] = []
        try:
            df = lookup_emails(ref_csv, config)
            emails_csv = str(runtime_path(paths.get("authors_emails_csv", "data/authors_emails.csv")))
            failed_csv = str(runtime_path(paths.get("failed_lookup_csv", "data/failed_lookup.csv")))
            save_results(df, emails_csv, failed_csv)
            task_status["progress"] = "完成"
        except Exception as e:
            task_status["progress"] = f"出错: {e}"
        finally:
            task_status["running"] = False

    thread = threading.Thread(target=run_in_background, daemon=True)
    thread.start()

    flash("邮箱检索已开始（后台运行中），请稍候刷新页面查看结果", "info")
    return redirect(url_for("lookup_page"))


@app.route("/lookup/upload-supplement", methods=["POST"])
def upload_supplement():
    """上传人工补充的邮箱 CSV"""
    config = get_runtime_config()
    paths = config.get("paths", {})

    f = request.files.get("supplement_file")
    if not f or not f.filename:
        flash("请选择文件", "warning")
        return redirect(url_for("lookup_page"))

    supp_csv = str(runtime_path(paths.get("manual_supplement_csv", "data/manual_supplement.csv")))
    runtime_path(paths["manual_supplement_csv"]).parent.mkdir(parents=True, exist_ok=True)
    f.save(supp_csv)

    emails_csv = str(runtime_path(paths.get("authors_emails_csv", "data/authors_emails.csv")))
    merge_manual_supplement(emails_csv, supp_csv)

    flash("人工补充数据已合并", "success")
    return redirect(url_for("lookup_page"))


# ---- 第三阶段：预览与发送 ----

@app.route("/send")
def send_page():
    config = get_runtime_config()
    paths = config.get("paths", {})

    emails_csv = runtime_path(paths.get("authors_emails_csv", "data/authors_emails.csv"))
    recipients = []
    if emails_csv.exists():
        try:
            recipients = prepare_email_data(str(emails_csv))
            repair_citing_authors_from_input(recipients, str(runtime_path(paths.get("input_dir", "data/input"))))
        except Exception:
            pass

    # 生成预览
    previews = []
    template_file = str(app_path(config.get("email", {}).get("template_file", "templates/notification.html")))
    journal_config = config.get("journal", {})

    # 加载邮件内容覆盖
    overrides_path = runtime_path(paths.get("email_overrides_json", "data/email_overrides.json"))
    html_overrides = {}
    if overrides_path.exists():
        try:
            with open(overrides_path, "r", encoding="utf-8") as f:
                html_overrides = json.load(f)
        except Exception:
            pass

    for r in recipients:
        try:
            subject, html_body = render_email(r, template_file, journal_config)
            if r["email"] in html_overrides:
                html_body = html_overrides[r["email"]]
            previews.append({
                "email": r["email"],
                "author_name": r["author_name"],
                "citation_count": len(r["citations"]),
                "subject": subject,
                "html_body": html_body,
            })
        except Exception as e:
            previews.append({
                "email": r["email"],
                "author_name": r["author_name"],
                "citation_count": len(r["citations"]),
                "subject": "渲染失败",
                "html_body": f"<p>模板渲染错误: {e}</p>",
            })

    # 已发送记录（每 timestamp+email 只展示一行）
    sent_data = []
    sent_csv = runtime_path(paths.get("sent_log_csv", "data/sent_log.csv"))
    if sent_csv.exists():
        try:
            df = load_sent_log(str(sent_csv))
            seen_keys = set()
            for s in df.to_dict("records"):
                key = (str(s.get("timestamp", ""))[:19], str(s.get("email", "")))
                if key not in seen_keys:
                    seen_keys.add(key)
                    sent_data.append(s)
        except Exception:
            pass

    return render_template(
        "send_page.html",
        previews=previews,
        sent_data=sent_data,
        task_status=task_status,
    )


@app.route("/send/run", methods=["POST"])
def run_send():
    if task_status["running"]:
        flash("已有任务正在运行，请等待完成", "warning")
        return redirect(url_for("send_page"))

    config = get_runtime_config()
    paths = config.get("paths", {})

    emails_csv = str(runtime_path(paths.get("authors_emails_csv", "data/authors_emails.csv")))
    if not os.path.exists(emails_csv):
        flash("没有可发送的邮箱数据", "warning")
        return redirect(url_for("send_page"))

    recipients = prepare_email_data(emails_csv)
    repair_citing_authors_from_input(recipients, str(runtime_path(paths.get("input_dir", "data/input"))))
    if not recipients:
        flash("没有可发送的收件人", "warning")
        return redirect(url_for("send_page"))

    selected = request.form.getlist("selected_emails[]")
    if selected:
        selected_set = set(selected)
        recipients = [r for r in recipients if r["email"] in selected_set]
        if not recipients:
            flash("所选收件人无有效数据", "warning")
            return redirect(url_for("send_page"))

    template_file = str(app_path(config.get("email", {}).get("template_file", "templates/notification.html")))
    journal_config = config.get("journal", {})
    smtp_config = config.get("smtp", {})
    email_config = config.get("email", {})
    sent_log_csv = str(runtime_path(paths.get("sent_log_csv", "data/sent_log.csv")))

    overrides_path = runtime_path(paths.get("email_overrides_json", "data/email_overrides.json"))
    html_overrides = {}
    if overrides_path.exists():
        try:
            with open(overrides_path, "r", encoding="utf-8") as f:
                html_overrides = json.load(f)
        except Exception:
            pass

    def run_in_background():
        task_status["running"] = True
        task_status["stage"] = "邮件发送"
        try:
            send_emails(
                recipients, template_file, journal_config,
                smtp_config, email_config, sent_log_csv,
                html_overrides=html_overrides,
                input_dir=str(runtime_path(paths["input_dir"])),
            )
            task_status["progress"] = "完成"
        except Exception as e:
            task_status["progress"] = f"出错: {e}"
        finally:
            task_status["running"] = False

    thread = threading.Thread(target=run_in_background, daemon=True)
    thread.start()

    flash(f"开始发送 {len(recipients)} 封邮件（后台运行中），请稍候刷新页面", "info")
    return redirect(url_for("send_page"))


@app.route("/api/task-status")
def api_task_status():
    return jsonify(task_status)


# ---- 无作者文献：加入检索 / 删除 ----

@app.route("/extract/promote-no-author", methods=["POST"])
def promote_no_author():
    """将无作者列表中的条目移入有作者列表（加入检索）"""
    config = get_runtime_config()
    paths = config.get("paths", {})
    no_author_csv = runtime_path(paths.get("other_references_csv", "data/other_references.csv"))
    ref_csv = runtime_path(paths.get("references_csv", "data/references.csv"))
    idx = request.form.get("idx", type=int)

    if idx is None or not no_author_csv.exists():
        flash("操作失败", "danger")
        return redirect(url_for("extract_page"))
    try:
        df_no = pd.read_csv(no_author_csv, encoding="utf-8-sig")
        row = df_no.iloc[[idx]]
        df_no = df_no.drop(index=idx).reset_index(drop=True)
        df_no.to_csv(no_author_csv, index=False, encoding="utf-8-sig")

        if ref_csv.exists():
            df_ref = pd.read_csv(ref_csv, encoding="utf-8-sig")
            df_ref = pd.concat([df_ref, row], ignore_index=True)
        else:
            df_ref = row
        df_ref.to_csv(ref_csv, index=False, encoding="utf-8-sig")

        flash("已加入检索列表", "success")
    except Exception as e:
        flash(f"操作失败: {e}", "danger")
    return redirect(url_for("extract_page"))


@app.route("/extract/promote-no-author-batch", methods=["POST"])
def promote_no_author_batch():
    """批量将无作者条目移入检索列表"""
    config = get_runtime_config()
    paths = config.get("paths", {})
    no_author_csv = runtime_path(paths.get("other_references_csv", "data/other_references.csv"))
    ref_csv = runtime_path(paths.get("references_csv", "data/references.csv"))
    indices = request.form.getlist("indices[]", type=int)

    if not indices or not no_author_csv.exists():
        flash("未选择任何条目", "warning")
        return redirect(url_for("extract_page"))
    try:
        df_no = pd.read_csv(no_author_csv, encoding="utf-8-sig")
        rows = df_no.iloc[indices]
        df_no = df_no.drop(index=indices).reset_index(drop=True)
        df_no.to_csv(no_author_csv, index=False, encoding="utf-8-sig")

        if ref_csv.exists():
            df_ref = pd.read_csv(ref_csv, encoding="utf-8-sig")
            df_ref = pd.concat([df_ref, rows], ignore_index=True)
        else:
            df_ref = rows
        df_ref.to_csv(ref_csv, index=False, encoding="utf-8-sig")

        flash(f"已将 {len(indices)} 条加入检索列表", "success")
    except Exception as e:
        flash(f"操作失败: {e}", "danger")
    return redirect(url_for("extract_page"))


@app.route("/extract/delete-no-author", methods=["POST"])
def delete_no_author_row():
    """从无作者列表中删除单条"""
    config = get_runtime_config()
    no_author_csv = runtime_path(config.get("paths", {}).get("other_references_csv", "data/other_references.csv"))
    idx = request.form.get("idx", type=int)
    if idx is None or not no_author_csv.exists():
        flash("删除失败", "danger")
        return redirect(url_for("extract_page"))
    try:
        df = pd.read_csv(no_author_csv, encoding="utf-8-sig")
        df = df.drop(index=idx).reset_index(drop=True)
        df.to_csv(no_author_csv, index=False, encoding="utf-8-sig")
        flash("已删除该条记录", "info")
    except Exception as e:
        flash(f"删除失败: {e}", "danger")
    return redirect(url_for("extract_page"))


@app.route("/extract/delete-no-author-batch", methods=["POST"])
def delete_no_author_batch():
    """批量删除无作者列表中的记录"""
    config = get_runtime_config()
    no_author_csv = runtime_path(config.get("paths", {}).get("other_references_csv", "data/other_references.csv"))
    indices = request.form.getlist("indices[]", type=int)
    if not indices or not no_author_csv.exists():
        flash("删除失败", "danger")
        return redirect(url_for("extract_page"))
    try:
        df = pd.read_csv(no_author_csv, encoding="utf-8-sig")
        df = df.drop(index=indices).reset_index(drop=True)
        df.to_csv(no_author_csv, index=False, encoding="utf-8-sig")
        flash(f"已删除 {len(indices)} 条记录", "info")
    except Exception as e:
        flash(f"删除失败: {e}", "danger")
    return redirect(url_for("extract_page"))


# ---- 级联删除辅助 ----

def _cascade_delete_by_titles(titles: list, config: dict):
    """从 authors_emails.csv 和 failed_lookup.csv 中删除 title 匹配的行"""
    paths = config.get("paths", {})
    for csv_key, default in [
        ("authors_emails_csv", "data/authors_emails.csv"),
        ("failed_lookup_csv", "data/failed_lookup.csv"),
    ]:
        csv_path = runtime_path(paths.get(csv_key, default))
        if not csv_path.exists():
            continue
        try:
            df = pd.read_csv(csv_path, encoding="utf-8-sig")
            if "title" in df.columns:
                df = df[~df["title"].astype(str).isin(titles)].reset_index(drop=True)
                df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        except Exception:
            pass


# ---- 删除单条数据 ----

@app.route("/extract/delete-ref", methods=["POST"])
def delete_ref_row():
    """删除单条参考文献记录"""
    config = get_runtime_config()
    ref_csv = runtime_path(config.get("paths", {}).get("references_csv", "data/references.csv"))
    idx = request.form.get("idx", type=int)
    if idx is None or not ref_csv.exists():
        flash("删除失败", "danger")
        return redirect(url_for("extract_page"))
    try:
        df = pd.read_csv(ref_csv, encoding="utf-8-sig")
        deleted_titles = [str(df.iloc[idx]["title"])] if "title" in df.columns else []
        df = df.drop(index=idx).reset_index(drop=True)
        df.to_csv(ref_csv, index=False, encoding="utf-8-sig")
        if deleted_titles:
            _cascade_delete_by_titles(deleted_titles, config)
        flash("已删除该条参考文献", "info")
    except Exception as e:
        flash(f"删除失败: {e}", "danger")
    return redirect(url_for("extract_page"))


@app.route("/lookup/delete-row", methods=["POST"])
def delete_lookup_row():
    """删除单条邮箱检索记录（已找到或未找到）"""
    config = get_runtime_config()
    paths = config.get("paths", {})
    csv_type = request.form.get("csv_type", "found")
    idx = request.form.get("idx", type=int)
    if csv_type == "found":
        target_csv = runtime_path(paths.get("authors_emails_csv", "data/authors_emails.csv"))
    else:
        target_csv = runtime_path(paths.get("failed_lookup_csv", "data/failed_lookup.csv"))
    if idx is None or not target_csv.exists():
        flash("删除失败", "danger")
        return redirect(url_for("lookup_page"))
    try:
        df = pd.read_csv(target_csv, encoding="utf-8-sig")
        df = df.drop(index=idx).reset_index(drop=True)
        df.to_csv(target_csv, index=False, encoding="utf-8-sig")
        flash("已删除该条记录", "info")
    except Exception as e:
        flash(f"删除失败: {e}", "danger")
    return redirect(url_for("lookup_page"))


@app.route("/lookup/edit-row", methods=["POST"])
def edit_lookup_row():
    """编辑检索结果行的 email / corresponding_author / matched_name"""
    config = get_runtime_config()
    paths = config.get("paths", {})
    emails_csv = runtime_path(paths.get("authors_emails_csv", "data/authors_emails.csv"))
    idx = request.form.get("idx", type=int)
    if idx is None or not emails_csv.exists():
        flash("编辑失败", "danger")
        return redirect(url_for("lookup_page"))
    try:
        df = pd.read_csv(emails_csv, encoding="utf-8-sig")
        if idx >= len(df):
            flash("索引超出范围", "danger")
            return redirect(url_for("lookup_page"))
        for field in ("email", "corresponding_author", "matched_name"):
            val = request.form.get(field)
            if val is not None and field in df.columns:
                df.at[idx, field] = val.strip()
        df.to_csv(emails_csv, index=False, encoding="utf-8-sig")
        flash("已更新该条记录", "success")
    except Exception as e:
        flash(f"编辑失败: {e}", "danger")
    return redirect(url_for("lookup_page"))


@app.route("/send/edit-recipient", methods=["POST"])
def edit_send_recipient():
    """编辑发送收件人（按原邮箱匹配，更新 email 和 matched_name）"""
    config = get_runtime_config()
    paths = config.get("paths", {})
    emails_csv = runtime_path(paths.get("authors_emails_csv", "data/authors_emails.csv"))
    original_email = request.form.get("original_email", "").strip()
    new_email = request.form.get("new_email", "").strip()
    new_author_name = request.form.get("new_author_name", "").strip()
    if not original_email or not emails_csv.exists():
        flash("编辑失败", "danger")
        return redirect(url_for("send_page"))
    try:
        df = pd.read_csv(emails_csv, encoding="utf-8-sig")
        mask = df["email"].astype(str).str.strip() == original_email
        if not mask.any():
            flash(f"未找到收件人: {original_email}", "warning")
            return redirect(url_for("send_page"))
        if new_email:
            df.loc[mask, "email"] = new_email
        if new_author_name and "matched_name" in df.columns:
            df.loc[mask, "matched_name"] = new_author_name
        df.to_csv(emails_csv, index=False, encoding="utf-8-sig")
        flash(f"已更新收件人: {original_email}", "success")
    except Exception as e:
        flash(f"编辑失败: {e}", "danger")
    return redirect(url_for("send_page"))


@app.route("/send/delete-recipient", methods=["POST"])
def delete_send_recipient():
    """从邮箱数据中删除某个收件人（按邮箱地址）"""
    config = get_runtime_config()
    paths = config.get("paths", {})
    email_to_del = request.form.get("email", "").strip()
    emails_csv = runtime_path(paths.get("authors_emails_csv", "data/authors_emails.csv"))
    if not email_to_del or not emails_csv.exists():
        flash("删除失败", "danger")
        return redirect(url_for("send_page"))
    try:
        df = pd.read_csv(emails_csv, encoding="utf-8-sig")
        df = df[df["email"].astype(str).str.strip() != email_to_del].reset_index(drop=True)
        df.to_csv(emails_csv, index=False, encoding="utf-8-sig")
        flash(f"已删除收件人: {email_to_del}", "info")
    except Exception as e:
        flash(f"删除失败: {e}", "danger")
    return redirect(url_for("send_page"))


# ---- 批量删除 ----

@app.route("/extract/delete-refs-batch", methods=["POST"])
def delete_refs_batch():
    config = get_runtime_config()
    ref_csv = runtime_path(config.get("paths", {}).get("references_csv", "data/references.csv"))
    indices = request.form.getlist("indices[]", type=int)
    if not indices or not ref_csv.exists():
        flash("未选择任何条目", "warning")
        return redirect(url_for("extract_page"))
    try:
        df = pd.read_csv(ref_csv, encoding="utf-8-sig")
        deleted_titles = [str(df.iloc[i]["title"]) for i in indices if i < len(df)] if "title" in df.columns else []
        df = df.drop(index=indices).reset_index(drop=True)
        df.to_csv(ref_csv, index=False, encoding="utf-8-sig")
        if deleted_titles:
            _cascade_delete_by_titles(deleted_titles, config)
        flash(f"已删除 {len(indices)} 条参考文献", "info")
    except Exception as e:
        flash(f"删除失败: {e}", "danger")
    return redirect(url_for("extract_page"))


@app.route("/lookup/delete-rows-batch", methods=["POST"])
def delete_lookup_rows_batch():
    config = get_runtime_config()
    paths = config.get("paths", {})
    csv_type = request.form.get("csv_type", "found")
    indices = request.form.getlist("indices[]", type=int)
    target_csv = runtime_path(paths.get(
        "authors_emails_csv" if csv_type == "found" else "failed_lookup_csv",
        "data/authors_emails.csv" if csv_type == "found" else "data/failed_lookup.csv",
    ))
    if not indices or not target_csv.exists():
        flash("未选择任何条目", "warning")
        return redirect(url_for("lookup_page"))
    try:
        df = pd.read_csv(target_csv, encoding="utf-8-sig")
        df = df.drop(index=indices).reset_index(drop=True)
        df.to_csv(target_csv, index=False, encoding="utf-8-sig")
        flash(f"已删除 {len(indices)} 条记录", "info")
    except Exception as e:
        flash(f"删除失败: {e}", "danger")
    return redirect(url_for("lookup_page"))


@app.route("/send/delete-recipients-batch", methods=["POST"])
def delete_send_recipients_batch():
    config = get_runtime_config()
    paths = config.get("paths", {})
    emails_to_del = request.form.getlist("emails[]")
    emails_csv = runtime_path(paths.get("authors_emails_csv", "data/authors_emails.csv"))
    if not emails_to_del or not emails_csv.exists():
        flash("未选择任何收件人", "warning")
        return redirect(url_for("send_page"))
    try:
        df = pd.read_csv(emails_csv, encoding="utf-8-sig")
        df = df[~df["email"].astype(str).str.strip().isin(emails_to_del)].reset_index(drop=True)
        df.to_csv(emails_csv, index=False, encoding="utf-8-sig")
        flash(f"已删除 {len(emails_to_del)} 位收件人", "info")
    except Exception as e:
        flash(f"删除失败: {e}", "danger")
    return redirect(url_for("send_page"))


@app.route("/lookup/supplement-email", methods=["POST"])
def supplement_email():
    """在未找到邮箱的条目上直接补充邮箱，移入已找到列表"""
    config = get_runtime_config()
    paths = config.get("paths", {})
    failed_csv = runtime_path(paths.get("failed_lookup_csv", "data/failed_lookup.csv"))
    emails_csv = runtime_path(paths.get("authors_emails_csv", "data/authors_emails.csv"))
    idx = request.form.get("idx", type=int)
    new_email = request.form.get("email", "").strip()
    if idx is None or not new_email or not failed_csv.exists():
        flash("补充失败：缺少必要参数", "danger")
        return redirect(url_for("lookup_page"))
    try:
        df_failed = pd.read_csv(failed_csv, encoding="utf-8-sig")
        if idx >= len(df_failed):
            flash("索引超出范围", "danger")
            return redirect(url_for("lookup_page"))
        row = df_failed.iloc[idx].to_dict()
        df_failed = df_failed.drop(index=idx).reset_index(drop=True)
        df_failed.to_csv(failed_csv, index=False, encoding="utf-8-sig")

        row["email"] = new_email
        row["lookup_source"] = "人工补充"
        row["matched_name"] = row.get("authors", "")
        new_row_df = pd.DataFrame([row])
        if emails_csv.exists():
            df_found = pd.read_csv(emails_csv, encoding="utf-8-sig")
            df_found = pd.concat([df_found, new_row_df], ignore_index=True)
        else:
            df_found = new_row_df
        df_found.to_csv(emails_csv, index=False, encoding="utf-8-sig")
        flash(f"已补充邮箱并移入已找到列表: {new_email}", "success")
    except Exception as e:
        flash(f"补充失败: {e}", "danger")
    return redirect(url_for("lookup_page"))


@app.route("/send/save-email-override", methods=["POST"])
def save_email_override():
    """保存手动编辑的邮件 HTML 内容"""
    email_addr = request.form.get("email", "").strip()
    html_body = request.form.get("html_body", "")
    if not email_addr:
        flash("邮箱地址不能为空", "warning")
        return redirect(url_for("send_page"))
    config = get_runtime_config()
    paths = config.get("paths", {})
    overrides_path = runtime_path(paths.get("email_overrides_json", "data/email_overrides.json"))
    try:
        overrides_path.parent.mkdir(parents=True, exist_ok=True)
        overrides = {}
        if overrides_path.exists():
            with open(overrides_path, "r", encoding="utf-8") as f:
                overrides = json.load(f)
        overrides[email_addr] = html_body
        with open(overrides_path, "w", encoding="utf-8") as f:
            json.dump(overrides, f, ensure_ascii=False, indent=2)
        flash(f"已保存邮件内容修改: {email_addr}", "success")
    except Exception as e:
        flash(f"保存失败: {e}", "danger")
    return redirect(url_for("send_page"))


@app.route("/send/export-log")
def export_sent_log():
    """导出发送日志为格式化 Excel"""
    from datetime import datetime as _dt

    config = get_runtime_config()
    paths = config.get("paths", {})
    sent_csv = runtime_path(paths.get("sent_log_csv", "data/sent_log.csv"))

    if not sent_csv.exists():
        flash("发送日志不存在", "warning")
        return redirect(url_for("send_page"))

    df = load_sent_log(str(sent_csv))

    COL_MAP = [
        ("发送时间",      "timestamp"),
        ("收件人邮箱",    "email"),
        ("作者姓名",      "author_name"),
        ("被引用文章",    "cited_paper_title"),
        ("引用了的文章",  "citing_paper_title"),
        ("发送状态",      "status"),
    ]
    COL_WIDTHS = [22, 32, 12, 40, 40, 14]

    header_fill   = PatternFill("solid", fgColor=get_active_journal()["primary"].lstrip("#").upper())
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    success_fill  = PatternFill("solid", fgColor="EAF4EA")
    fail_fill     = PatternFill("solid", fgColor="FDECEA")
    fail_font     = Font(color="CC0000")

    wb = Workbook()
    ws = wb.active
    ws.title = "发送记录"

    for col_idx, (header, _) in enumerate(COL_MAP, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(df.to_dict("records"), 2):
        is_success = str(row.get("status", "")) == "成功"
        row_fill = success_fill if is_success else fail_fill
        for col_idx, (_, csv_col) in enumerate(COL_MAP, 1):
            val = str(row.get(csv_col, ""))
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if not is_success and col_idx == len(COL_MAP):
                cell.font = fail_font

    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"引用通知发送记录_{_dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---- 数据下载 ----

@app.route("/download/<filename>")
def download_file(filename):
    safe_names = {
        "references.csv", "other_references.csv", "authors_emails.csv",
        "failed_lookup.csv", "sent_log.csv", "manual_supplement.csv",
    }
    if filename not in safe_names:
        flash("无效的文件名", "danger")
        return redirect(url_for("index"))

    config = get_runtime_config()
    paths = config.get("paths", {})
    data_dir = runtime_path(paths["references_csv"]).parent
    filepath = data_dir / filename
    if not filepath.exists():
        flash(f"文件不存在: {filename}", "warning")
        return redirect(url_for("index"))

    return send_from_directory(str(data_dir), filename, as_attachment=True)


# ==================== 启动 ====================

if __name__ == "__main__":
    for journal in JOURNALS.values():
        config = load_config()
        paths = build_journal_paths(journal, config)
        ensure_journal_dirs(paths)
    print("=" * 50)
    print("  多刊物引用通知系统")
    print("  请在浏览器中打开: http://127.0.0.1:5000")
    print("=" * 50)
    app.run(debug=True, host="127.0.0.1", port=5000)

